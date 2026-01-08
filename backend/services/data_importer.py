"""
Data Importer Service

Handles importing data from various sources (Excel, CSV, images with barcodes, etc.)
Parses files and returns data as dictionaries for the DAL layer to insert.

Usage:
    from services.data_importer import import_products_from_excel
    docs = import_products_from_excel("data.xlsx", sheet_name="products")
"""

from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional
import json
from openpyxl import load_workbook
import pandas as pd


def import_products_from_excel(file_path: str, sheet_name: str = "products") -> List[Dict[str, Any]]:
    """
    Parse products from an Excel file.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (default: "products")
        
    Returns:
        List of product documents ready for database insertion
    """
    handler = ExcelHandler()
    return handler.read(file_path, sheet_name)


def import_products_from_csv(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse products from a CSV file.
    
    Args:
        file_path: Path to the CSV file
        
    Returns:
        List of product documents ready for database insertion
    """
    handler = CSVHandler()
    return handler.read(file_path)


class DataImporter:
    """Main class for importing data from multiple sources."""

    def __init__(self):
        """Initialize MongoDB connection."""
        load_dotenv()
        self.mongo_uri = os.getenv(
            "MONGO_URI", "mongodb://admin:password@localhost:27017/?authSource=admin"
        )
        self.db_name = os.getenv("MONGO_INITDB_DATABASE", "retail_db")
        self.client = None
        self.db = None

    def connect_mongodb(self) -> bool:
        """
        Connect to MongoDB.
        
        Returns:
            bool: True if connection successful, False otherwise
        """
        try:
            self.client = MongoClient(self.mongo_uri)
            self.db = self.client[self.db_name]
            # Test connection
            self.client.admin.command('ping')
            return True
        except Exception as e:
            print(f"\n✗ Error connecting to MongoDB: {e}")
            print("Make sure MongoDB is running (docker compose up -d)")
            return False    

    def disconnect_mongodb(self):
        """Disconnect from MongoDB."""
        if self.client:
            self.client.close()
            self.client = None
            self.db = None

    def import_from_excel(
        self, file_path: str, sheet_name: str = "Sheet1"
    ) -> List[Dict[str, Any]]:
        """
        Import data from Excel file using ExcelHandler.
        
        Args:
            file_path (str): Path to Excel file
            sheet_name (str): Name of sheet to read
            
        Returns:
            List[Dict]: List of documents ready for MongoDB
        """
        handler = ExcelHandler()
        data = handler.read(file_path, sheet_name)
        
        # Print JSON output for verification
        print(json.dumps(data, indent=2, ensure_ascii=False, default=str))
        return data
       

    def import_from_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Import data from CSV file.
        
        Args:
            file_path (str): Path to CSV file
            
        Returns:
            List[Dict]: List of documents ready for MongoDB
        """
        handler=CSVHandler()
        data=handler.read(file_path=file_path)
        print(json.dumps(data, indent=2, ensure_ascii=False, default=str))
        return data

    def import_from_image_barcode(
        self, image_path: str, barcode_type: str = "auto"
    ) -> Dict[str, Any]:
        """
        Import data from image with barcode.
        Extracts barcode and returns as document.
        
        Args:
            image_path (str): Path to image file
            barcode_type (str): Type of barcode (auto, code128, ean13, etc.)
            
        Returns:
            Dict: Document with barcode data
        """
        # TODO: Implement barcode scanning from image
        pass

    def import_from_barcode_text(self, barcode: str) -> Dict[str, Any]:
        """
        Import/lookup data from barcode text.
        
        Args:
            barcode (str): Barcode string
            
        Returns:
            Dict: Document with product info
        """
        # TODO: Implement barcode text parsing
        pass


    def save_to_mongodb(
        self,
        data: List[Dict[str, Any]],
        collection: str = "products",
        upsert: bool = False,
    ) -> Dict[str, Any]:
        """
        Save data to MongoDB collection.
        
        Args:
            data (List[Dict]): Documents to insert
            collection (str): MongoDB collection name
            upsert (bool): If True, update existing docs; if False, insert new
            
        Returns:
            Dict: Result with inserted_ids, modified_count, etc.
        """
        if not self.client or not self.db:
            raise Exception("Not connected to MongoDB. Call connect_mongodb() first.")
        
        coll = self.db[collection]  # Use the collection parameter
        
        # Ensure data is a list
        documents = data if isinstance(data, list) else list(data.values())
            
        # Insert into MongoDB
        if documents:
            result = coll.insert_many(documents)
            print(f"\n✓ Successfully inserted {len(result.inserted_ids)} documents into '{self.db_name}.{collection}'")
            return {
                "inserted_count": len(result.inserted_ids),
                "inserted_ids": [str(id) for id in result.inserted_ids],
            }
        else:
            print("\n⚠ No documents to insert")
            return {"inserted_count": 0, "inserted_ids": []}

    def get_import_stats(self) -> Dict[str, Any]:
        """
        Get statistics about imports.
        
        Returns:
            Dict: Stats like total imported, by source, errors, etc.
        """
        # TODO: Implement stats tracking
        pass


class SourceHandler(ABC):
    """Abstract base class for different import sources."""

    @abstractmethod
    def read(self, source: str) -> List[Dict[str, Any]]:
        """Read data from source."""
        pass

    @abstractmethod
    def validate(self, data: List[Dict[str, Any]]) -> tuple[bool, List[str]]:
        """Validate source-specific data format."""
        pass


class ExcelHandler(SourceHandler):
    """Handles Excel file imports."""

    def read(self, file_path: str, sheet_name: str = "Sheet1") -> List[Dict[str, Any]]:
        """
        Read Excel file.
        
        Args:
            file_path (str): Path to Excel file
            sheet_name (str): Sheet to read
            
        Returns:
            List[Dict]: Documents from Excel
        """
        workbook = load_workbook(filename=file_path)
        
        # Handle sheet naming: try requested name, fallback to default names
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        elif "Sheet" in workbook.sheetnames:
            sheet = workbook["Sheet"]
        else:
            sheet = workbook.active
        
        documents = []
        fields = []

        # Read headers from first row (all columns)
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            fields = list(row)  # Get all column headers

        # Read data rows (all columns to match headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Build document dict dynamically based on actual columns
            document = {}
            for i, field in enumerate(fields):
                if i < len(row) and field:  # Only add if field name exists and value is present
                    document[field] = row[i]
            
            if document:  # Only add non-empty documents
                documents.append(document)
        
        return documents

    def validate(self, data: List[Dict[str, Any]]) -> tuple[bool, List[str]]:
        """Validate Excel data."""
        errors = []
        
        if not data:
            errors.append("No data to validate")
            return False, errors
        
        # Check that all documents have at least one field
        for i, doc in enumerate(data):
            if not doc:
                errors.append(f"Row {i+2}: Empty document")
        
        is_valid = len(errors) == 0
        return is_valid, errors


class CSVHandler(SourceHandler):
    """Handles CSV file imports."""

    def read(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Read CSV file.
        
        Args:
            file_path (str): Path to CSV file
            
        Returns:
            List[Dict]: Documents from CSV
        """
        try:
            # Read CSV with pandas
            df = pd.read_csv(file_path)
            
            # Convert to list of dictionaries (row-oriented)
            documents = df.to_dict(orient='records')
            
            # Remove rows with all NaN values
            documents = [doc for doc in documents if any(doc.values())]
            
            return documents
        
        except FileNotFoundError:
            print(f"✗ CSV file not found: {file_path}")
            return []
        except Exception as e:
            print(f"✗ Error reading CSV file: {e}")
            return []

    def validate(self, data: List[Dict[str, Any]]) -> tuple[bool, List[str]]:
        """Validate CSV data."""
        errors = []
        
        if not data:
            errors.append("No data to validate")
            return False, errors
        
        # Check that all documents have at least one field
        for i, doc in enumerate(data):
            if not doc or all(v is None or (isinstance(v, float) and pd.isna(v)) for v in doc.values()):
                errors.append(f"Row {i+1}: Empty or all-null document")
        
        is_valid = len(errors) == 0
        return is_valid, errors


class BarcodeHandler(SourceHandler):
    """Handles barcode scanning and lookup."""

    def read(self, barcode: str) -> List[Dict[str, Any]]:
        """
        Process barcode.
        
        Args:
            barcode (str): Barcode string or image path
            
        Returns:
            List[Dict]: Product data associated with barcode
        """
        # TODO: Implement barcode processing
        pass

    def validate(self, data: List[Dict[str, Any]]) -> tuple[bool, List[str]]:
        """Validate barcode data."""
        # TODO: Implement validation
        pass
