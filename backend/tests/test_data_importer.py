"""
Tests for DataImporter service (Excel and CSV imports).

Run with:
  pytest tests/test_data_importer.py -v
  
Or without pytest:
  python -m tests.test_data_importer
"""

import unittest
import os
import tempfile
from pathlib import Path
from openpyxl import Workbook
import csv

from app.services.data_importer import DataImporter, ExcelHandler, CSVHandler


class TestExcelHandler(unittest.TestCase):
    """Test cases for Excel file imports."""

    def setUp(self):
        """Create temporary Excel file for testing."""
        self.temp_dir = tempfile.mkdtemp()
        self.excel_file = os.path.join(self.temp_dir, "test_products.xlsx")
        
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "products"
        
        # Headers
        ws.append(["sku", "name", "category", "price"])
        
        # Data rows
        ws.append(["SKU-001", "Product A", "electronics", 99.99])
        ws.append(["SKU-002", "Product B", "clothing", 29.99])
        ws.append(["SKU-003", "Product C", "food", 5.99])
        ws.append(["SKU-004", "Product D", "home", 15.50])
        
        wb.save(self.excel_file)

    def tearDown(self):
        """Clean up temporary files."""
        if os.path.exists(self.excel_file):
            os.remove(self.excel_file)
        os.rmdir(self.temp_dir)

    def test_excel_read_basic(self):
        """Test basic Excel file reading."""
        handler = ExcelHandler()
        data = handler.read(self.excel_file, sheet_name="products")
        
        # Check count
        self.assertEqual(len(data), 4, "Should read 4 product rows")
        
        # Check first document
        first = data[0]
        self.assertEqual(first["sku"], "SKU-001")
        self.assertEqual(first["name"], "Product A")
        self.assertEqual(first["category"], "electronics")
        self.assertAlmostEqual(first["price"], 99.99)

    def test_excel_read_all_fields(self):
        """Test that all fields are read correctly."""
        handler = ExcelHandler()
        data = handler.read(self.excel_file, sheet_name="products")
        
        # Check all products have required fields
        for product in data:
            self.assertIn("sku", product)
            self.assertIn("name", product)
            self.assertIn("category", product)
            self.assertIn("price", product)

    def test_excel_validate_success(self):
        """Test validation of valid Excel data."""
        handler = ExcelHandler()
        data = handler.read(self.excel_file, sheet_name="products")
        
        is_valid, errors = handler.validate(data)
        self.assertTrue(is_valid, "Valid data should pass validation")
        self.assertEqual(len(errors), 0, "Should have no validation errors")

    def test_excel_read_returns_list(self):
        """Test that Excel handler returns a list, not a dict."""
        handler = ExcelHandler()
        data = handler.read(self.excel_file, sheet_name="products")
        
        self.assertIsInstance(data, list, "Should return a list")
        self.assertTrue(len(data) > 0, "List should not be empty")
        self.assertIsInstance(data[0], dict, "Each item should be a dict")


class TestCSVHandler(unittest.TestCase):
    """Test cases for CSV file imports."""

    def setUp(self):
        """Create temporary CSV file for testing."""
        self.temp_dir = tempfile.mkdtemp()
        self.csv_file = os.path.join(self.temp_dir, "test_products.csv")
        
        # Create test CSV file
        with open(self.csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["sku", "name", "category", "price"])
            writer.writerow(["SKU-001", "Product A", "electronics", "99.99"])
            writer.writerow(["SKU-002", "Product B", "clothing", "29.99"])
            writer.writerow(["SKU-003", "Product C", "food", "5.99"])
            writer.writerow(["SKU-004", "Product D", "home", "15.50"])

    def tearDown(self):
        """Clean up temporary files."""
        if os.path.exists(self.csv_file):
            os.remove(self.csv_file)
        os.rmdir(self.temp_dir)

    def test_csv_read_basic(self):
        """Test basic CSV file reading."""
        handler = CSVHandler()
        data = handler.read(self.csv_file)
        
        # Check count
        self.assertEqual(len(data), 4, "Should read 4 product rows")
        
        # Check first document
        first = data[0]
        self.assertEqual(first["sku"], "SKU-001")
        self.assertEqual(first["name"], "Product A")
        self.assertEqual(first["category"], "electronics")
        # CSV reads as string, not float
        self.assertEqual(str(first["price"]), "99.99")

    def test_csv_read_all_fields(self):
        """Test that all fields are read correctly from CSV."""
        handler = CSVHandler()
        data = handler.read(self.csv_file)
        
        # Check all products have required fields
        for product in data:
            self.assertIn("sku", product)
            self.assertIn("name", product)
            self.assertIn("category", product)
            self.assertIn("price", product)

    def test_csv_validate_success(self):
        """Test validation of valid CSV data."""
        handler = CSVHandler()
        data = handler.read(self.csv_file)
        
        is_valid, errors = handler.validate(data)
        self.assertTrue(is_valid, "Valid data should pass validation")
        self.assertEqual(len(errors), 0, "Should have no validation errors")

    def test_csv_read_returns_list(self):
        """Test that CSV handler returns a list."""
        handler = CSVHandler()
        data = handler.read(self.csv_file)
        
        self.assertIsInstance(data, list, "Should return a list")
        self.assertTrue(len(data) > 0, "List should not be empty")
        self.assertIsInstance(data[0], dict, "Each item should be a dict")

    def test_csv_file_not_found(self):
        """Test handling of missing CSV file."""
        handler = CSVHandler()
        data = handler.read("/nonexistent/file.csv")
        
        self.assertEqual(data, [], "Should return empty list for missing file")


class TestDataImporter(unittest.TestCase):
    """Test cases for main DataImporter class."""

    def setUp(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()
        self.excel_file = os.path.join(self.temp_dir, "test_products.xlsx")
        self.csv_file = os.path.join(self.temp_dir, "test_products.csv")
        
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        # Note: default sheet name is 'Sheet', not 'Sheet1'
        ws.append(["sku", "name", "category", "price"])
        ws.append(["SKU-001", "Product A", "electronics", 99.99])
        ws.append(["SKU-002", "Product B", "clothing", 29.99])
        wb.save(self.excel_file)
        
        # Create test CSV file
        with open(self.csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["sku", "name", "category", "price"])
            writer.writerow(["SKU-101", "Product X", "electronics", "199.99"])
            writer.writerow(["SKU-102", "Product Y", "clothing", "49.99"])

    def tearDown(self):
        """Clean up temporary files."""
        for f in [self.excel_file, self.csv_file]:
            if os.path.exists(f):
                os.remove(f)
        os.rmdir(self.temp_dir)

    def test_import_from_excel(self):
        """Test DataImporter.import_from_excel()."""
        importer = DataImporter()
        data = importer.import_from_excel(self.excel_file)
        
        self.assertEqual(len(data), 2, "Should import 2 products from Excel")
        self.assertEqual(data[0]["sku"], "SKU-001")

    def test_import_from_csv(self):
        """Test DataImporter.import_from_csv()."""
        importer = DataImporter()
        data = importer.import_from_csv(self.csv_file)
        
        self.assertEqual(len(data), 2, "Should import 2 products from CSV")
        self.assertEqual(data[0]["sku"], "SKU-101")

    def test_import_methods_return_list(self):
        """Test that import methods return list of dicts."""
        importer = DataImporter()
        
        excel_data = importer.import_from_excel(self.excel_file)
        csv_data = importer.import_from_csv(self.csv_file)
        
        self.assertIsInstance(excel_data, list)
        self.assertIsInstance(csv_data, list)
        self.assertTrue(len(excel_data) > 0)
        self.assertTrue(len(csv_data) > 0)


class TestDataComparisonExcelVsCSV(unittest.TestCase):
    """Test cases comparing Excel and CSV imports."""

    def setUp(self):
        """Create identical Excel and CSV files."""
        self.temp_dir = tempfile.mkdtemp()
        self.excel_file = os.path.join(self.temp_dir, "products.xlsx")
        self.csv_file = os.path.join(self.temp_dir, "products.csv")
        
        # Create identical test data in both formats
        test_data = [
            ["sku", "name", "category", "price"],
            ["SKU-001", "Product A", "electronics", 99.99],
            ["SKU-002", "Product B", "clothing", 29.99],
            ["SKU-003", "Product C", "food", 5.99],
        ]
        
        # Excel version - use correct sheet naming
        wb = Workbook()
        ws = wb.active
        for row in test_data:
            ws.append(row)
        wb.save(self.excel_file)
        
        # CSV version
        with open(self.csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(test_data)

    def tearDown(self):
        """Clean up temporary files."""
        for f in [self.excel_file, self.csv_file]:
            if os.path.exists(f):
                os.remove(f)
        os.rmdir(self.temp_dir)

    def test_excel_and_csv_same_count(self):
        """Test that Excel and CSV imports produce same number of records."""
        importer = DataImporter()
        excel_data = importer.import_from_excel(self.excel_file)
        csv_data = importer.import_from_csv(self.csv_file)
        
        self.assertEqual(len(excel_data), len(csv_data), 
                        "Excel and CSV should import same number of records")

    def test_excel_and_csv_same_fields(self):
        """Test that Excel and CSV have same field names."""
        importer = DataImporter()
        excel_data = importer.import_from_excel(self.excel_file)
        csv_data = importer.import_from_csv(self.csv_file)
        
        excel_fields = set(excel_data[0].keys())
        csv_fields = set(csv_data[0].keys())
        
        self.assertEqual(excel_fields, csv_fields, 
                        "Excel and CSV should have same field names")


if __name__ == "__main__":
    unittest.main()
